"""Read station metadata from the public Zenodo workbooks."""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class StationMeta:
    station_id: str
    latitude: float
    longitude: float
    land_cover: str
    network: str


def load_station_metadata(path: str) -> dict[str, StationMeta]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        header = next(rows)
        index = {str(value): position for position, value in enumerate(header) if value is not None}
        required = ("Latitude", "Longitude", "General classification", "Data source/network")
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"station metadata missing columns: {', '.join(missing)}")
        result: dict[str, StationMeta] = {}
        for row in rows:
            station = row[0]
            if station is None or not str(station).strip():
                continue
            station_id = str(station).strip()
            result[station_id] = StationMeta(
                station_id=station_id,
                latitude=float(row[index["Latitude"]]),
                longitude=float(row[index["Longitude"]]),
                land_cover=str(row[index["General classification"]] or "").strip(),
                network=str(row[index["Data source/network"]] or "").strip(),
            )
        return result
    finally:
        workbook.close()
